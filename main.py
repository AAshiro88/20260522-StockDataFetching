import sys
import time
import threading
import requests
import random
import json
import os
import csv
from datetime import datetime, timezone, timedelta

# --- 引入 GUI 相關元件 ------------------------
import tkinter as tk
from tkinter import font as tkfont

# --- 本地 Excel 設定檔（唯讀，程式永不寫入）---

import openpyxl

# --- 時間設定 --------------------------------

TPE = timezone(timedelta(hours=8))

def now_tpe():
    return datetime.fromtimestamp(time.time(), TPE)

# --- 檔案設定 --------------------------------

if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

MAIN_XLSX = os.path.join(BASE_DIR, "main.xlsx")
HISTORY_CSV_DIR = os.path.join(BASE_DIR, "個股股價")
CSV_FILE = os.path.join(BASE_DIR, "即時報價.csv")

# --- 股票與 ETF 清單（從 main.xlsx 唯讀載入）---

WATCH_LIST: list[str] = []
ETF_CODES: set[str] = set()
HIGHLIGHT_STOCKS: list[str] = []

# --- 指數清單 --------------------------------

# 修正：櫃買指數代碼由 t01 修正為 o00，以對應正確的櫃買指數
INDEX_LIST = [
    {"code": "t00", "name": "加權指數", "sheet": "加權指數"},
    {"code": "o00", "name": "櫃買指數", "sheet": "櫃買指數"},
]

# --- API ──────────────────────────────────────

URL = "https://mis.twse.com.tw/stock/api/getStockInfo.jsp"
HEADERS = {"User-Agent": "Mozilla/5.0"}


# --- 輔助函式 --------------------------------

def safe_float(v):
    try:
        if isinstance(v, (list, dict)):
            return 0
        if v in ["-", "", None]:
            return 0
        return float(v)
    except Exception:
        return 0

def price_color_tag(v):
    try:
        v = float(v)
    except Exception:
        return "white"
    if v > 0:
        return "red"
    if v < 0:
        return "green"
    return "white"

# 修正：移除未被呼叫的 pad_stock_label 死碼函式


# --- 狀態管理類別 ----------------------------

class StockState:

    def __init__(self):
        self.data: dict = {}
        self.local_time: str = ""
        self.twse_time: str | None = None
        self.error: str = ""
        self.bad_codes: list[str] = []

        self._price_cache: dict[str, float] = {}
        self._y_cache: dict[str, float] = {}
        self._history_row_cache: dict[str, list] = {}
        self._sheet_lock = threading.Lock()
        self._sheet_writing = False
        self._csv_lock = threading.Lock()
        self._csv_writing = False
        self._prev_tick_price: dict[str, float] = {}
        self._alert_time: dict[str, float] = {}
        # 儲存警示觸發時的漲跌方向，用於區分顏色
        self._alert_dir: dict[str, str] = {}
        # 保護 state.data 讀寫的鎖，所有執行緒讀取前必須先取得此鎖的 snapshot
        self._data_lock = threading.Lock()

        self._industry_cache: dict[str, str] = {}

        self._init_data_files()
        self._load_main_config()
        self._load_industry_mapping()

    def _init_data_files(self):
        """初始化資料檔（唯讀，永不寫入現有檔案）"""
        os.makedirs(HISTORY_CSV_DIR, exist_ok=True)
        if not os.path.exists(MAIN_XLSX):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "main"
            ws.append(["股號", "ETF代碼", "高亮股號"])
            wb.save(MAIN_XLSX)
            wb.close()

    def _load_main_config(self):
        """從 main.xlsx 唯讀載入設定（可用 Excel 編輯 main.xlsx）"""
        global WATCH_LIST, ETF_CODES, HIGHLIGHT_STOCKS
        try:
            if not os.path.exists(MAIN_XLSX):
                return
            wb = openpyxl.load_workbook(MAIN_XLSX, read_only=True)
            if "main" not in wb.sheetnames:
                wb.close()
                return
            ws = wb["main"]
            watch = []
            etf = set()
            highlight = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) >= 1 and row[0] is not None:
                    val = str(row[0]).strip()
                    if val:
                        watch.append(val)
                if len(row) >= 2 and row[1] is not None:
                    val = str(row[1]).strip()
                    if val:
                        etf.add(val)
                if len(row) >= 3 and row[2] is not None:
                    val = str(row[2]).strip()
                    if val:
                        highlight.append(val)
            wb.close()
            WATCH_LIST = watch
            ETF_CODES = etf
            HIGHLIGHT_STOCKS = highlight
        except Exception as e:
            self.error = f"載入 main 設定表失敗: {e}"

    def _load_industry_mapping(self):
        """從 main.xlsx 讀取產業別對照表"""
        try:
            if not os.path.exists(MAIN_XLSX):
                return
            wb = openpyxl.load_workbook(MAIN_XLSX, read_only=True)
            sheet_name = None
            for name in wb.sheetnames:
                if "一覽表" in name:
                    sheet_name = name
                    break
            if not sheet_name:
                wb.close()
                return
            ws = wb[sheet_name]
            all_values = []
            for row in ws.iter_rows(values_only=True):
                all_values.append(list(row))
            wb.close()
            if len(all_values) <= 4:
                return
            header = all_values[4]
            code_idx, industry_idx = -1, -1
            for idx, name in enumerate(header):
                if name and "有價證券代號及名稱" in str(name):
                    code_idx = idx
                elif name and "產業別" in str(name):
                    industry_idx = idx
            if code_idx == -1: code_idx = 0
            if industry_idx == -1: industry_idx = 4
            for row in all_values[5:]:
                if len(row) <= max(code_idx, industry_idx):
                    continue
                raw = str(row[code_idx]).strip() if row[code_idx] else ""
                ind = str(row[industry_idx]).strip() if row[industry_idx] else ""
                if raw and ind:
                    parts = raw.replace("\u3000", " ").split()
                    if parts:
                        code = parts[0].strip()
                        if code.isdigit():
                            self._industry_cache[code] = ind
        except Exception as e:
            self.error = f"載入產業別對照表失敗: {e}"

    def get_industry(self, code: str) -> str:
        return self._industry_cache.get(code, "其他業")




# --- 資料抓取 --------------------------------

# 修正：將 parse_five 移至模組層級，避免在迴圈內重複建立 function物件
def parse_five(raw: str) -> list:
    """解析以底線分隔的五檔報價字串，回傳長度為 5 的 float 清單"""
    if not raw or raw == "-":
        return [0, 0, 0, 0, 0]
    parts = raw.split("_")
    values = []
    for p in parts:
        if len(values) >= 5:
            break
        if p != "":
            values.append(safe_float(p))
    while len(values) < 5:
        values.append(0)
    return values


def fetch(state: StockState) -> tuple[dict, list[str]]:
    result = {}
    twse_times = []

    combined_watch_list = list(set(WATCH_LIST).union(ETF_CODES))

    # 修正：動態判斷指數代碼，o00 開頭使用 otc_ 市場別，其餘使用 tse_ 市場別
    index_ex_list = []
    for i in INDEX_LIST:
        prefix = "otc" if i["code"].startswith("o") else "tse"
        index_ex_list.append(f"{prefix}_{i['code']}.tw")

    batch_size = 50
    batches = [combined_watch_list[i:i+batch_size]
               for i in range(0, len(combined_watch_list), batch_size)]

    try:
        for batch_idx, batch in enumerate(batches):
            time.sleep(random.uniform(0.2, 0.5))

            ex = "|".join(
                [f"tse_{c}.tw|otc_{c}.tw" for c in batch]
                + (index_ex_list if batch_idx == 0 else [])
            )

            r = requests.get(
                URL,
                params={"ex_ch": ex, "json": 1, "delay": 0},
                headers=HEADERS,
                timeout=10,
            )

            if r.status_code != 200:
                state.error = f"抓取失敗: 證交所拒絕連線 (HTTP {r.status_code})"
                return result, twse_times

            try:
                data = r.json()
            except json.JSONDecodeError:
                state.error = "抓取失敗: 證交所未回傳標準資料 (可能受到暫時性限制)"
                return result, twse_times

            state.error = ""

            for i in data.get("msgArray", []):
                c = i.get("c")
                if not c or c in result:
                    continue

                prev = safe_float(i.get("y"))

                if prev > 0:
                    state._y_cache[c] = prev
                elif c in state._y_cache:
                    prev = state._y_cache[c]

                price = safe_float(i.get("z"))

                if price == 0:
                    price = safe_float(i.get("tv"))

                open_p = safe_float(i.get("o"))
                high_p = safe_float(i.get("h"))
                low_p  = safe_float(i.get("l"))

                if price == 0 and c in state._price_cache:
                    price = state._price_cache[c]

                if price > 0:
                    state._price_cache[c] = price

                if prev == 0 and c in state._y_cache:
                    prev = state._y_cache[c]

                if price > 0 and prev > 0:
                    chg = round(price - prev, 2)
                    pct = round(chg / prev * 100, 2) if prev else 0
                else:
                    chg = 0
                    pct = 0

                if high_p == 0 and price > 0:
                    high_p = price
                if low_p == 0 and price > 0:
                    low_p = price

                t = i.get("t")
                if t:
                    twse_times.append(t)

                prev_close = prev
                volume     = safe_float(i.get("v"))

                if prev_close > 0 and high_p > 0 and low_p > 0:
                    amplitude = round((high_p - low_p) / prev_close * 100, 2)
                else:
                    amplitude = 0

                bid_prices = parse_five(i.get("b", ""))
                bid_vols   = parse_five(i.get("g", ""))
                ask_prices = parse_five(i.get("f", ""))
                ask_vols   = parse_five(i.get("g", ""))

                prev_tick = state._prev_tick_price.get(c, 0)
                if prev_tick > 0 and price > 0:
                    tick_chg_pct = (price - prev_tick) / prev_tick * 100
                    abs_tick_chg_pct = abs(tick_chg_pct)
                else:
                    tick_chg_pct = 0
                    abs_tick_chg_pct = 0

                if price > 0:
                    state._prev_tick_price[c] = price

                now_ts = time.time()
                if abs_tick_chg_pct >= 1:
                    state._alert_time[c] = now_ts
                    state._alert_dir[c] = "up" if tick_chg_pct > 0 else "down"
                elif c in state._alert_time:
                    if now_ts - state._alert_time[c] >= 300:
                        del state._alert_time[c]
                        if c in state._alert_dir:
                            del state._alert_dir[c]

                result[c] = {
                    "name":       i.get("n", ""),
                    "price":      price,
                    "prev_close": prev_close,
                    "chg":        chg,
                    "pct":        pct,
                    "open":       open_p,
                    "high":       high_p,
                    "low":        low_p,
                    "volume":     volume,
                    "amplitude":  amplitude,
                    "alert":      c in state._alert_time,
                    "alert_dir":  state._alert_dir.get(c, ""),
                    "bid_prices": bid_prices,
                    "bid_vols":   bid_vols,
                    "ask_prices": ask_prices,
                    "ask_vols":   ask_vols,
                    "time":       t,
                    "code":       c,
                }

        state.bad_codes = [
            c for c in WATCH_LIST
            if c not in result and c not in ETF_CODES
        ]

    except Exception as e:
        state.error = f"抓取失敗: {e}"

    return result, twse_times


def refresh(state: StockState) -> None:
    data, twse_times = fetch(state)
    state.local_time = now_tpe().strftime("%H:%M:%S")

    if data:
        # 以鎖保護寫入，確保讀取端不會取到被部分覆寫的中間狀態
        with state._data_lock:
            state.data = data
            if twse_times:
                state.twse_time = twse_times[0]


# --- 歷史資料寫入（CSV 安全寫入）-------------

def push_to_sheet(state: StockState):
    today = now_tpe().strftime("%Y-%m-%d")

    with state._data_lock:
        data_snapshot = dict(state.data)

    os.makedirs(HISTORY_CSV_DIR, exist_ok=True)

    combined_watch_list = list(set(WATCH_LIST).union(ETF_CODES))
    all_codes = combined_watch_list + [i["code"] for i in INDEX_LIST]

    for c in all_codes:
        try:
            d = data_snapshot.get(c)
            if not d or not d.get("time"):
                continue
            if d["price"] == 0:
                continue

            csv_path = os.path.join(HISTORY_CSV_DIR, f"{c}.csv")
            header = ["日期", "開盤價", "最高價", "最低價", "收盤價", "漲跌", "漲跌幅(%)"]

            rows = []
            last_date = None
            if os.path.exists(csv_path):
                with open(csv_path, newline="", encoding="utf-8-sig") as f:
                    reader = csv.reader(f)
                    next(reader, None)
                    for row in reader:
                        if row:
                            rows.append(row)
                            last_date = row[0] if row[0] else None

            target_open = d["open"]
            target_high = d["high"]
            target_low  = d["low"]

            if last_date == today and rows:
                last_row = rows[-1]
                try:
                    ext_open = float(last_row[1]) if last_row[1] else 0
                    ext_high = float(last_row[2]) if last_row[2] else 0
                    ext_low  = float(last_row[3]) if last_row[3] else 0
                    if ext_open > 0:
                        target_open = ext_open
                    if ext_high > 0:
                        target_high = max(ext_high, target_high)
                    if target_low > 0 and ext_low > 0:
                        target_low = min(ext_low, target_low)
                    elif ext_low > 0:
                        target_low = ext_low
                except (ValueError, IndexError):
                    pass

            new_row = [today, target_open, target_high, target_low, d["price"], d["chg"], d["pct"]]

            cache_key = f"hist_code_{c}"
            if state._history_row_cache.get(cache_key) == new_row:
                continue

            if last_date == today and rows:
                rows[-1] = new_row
            else:
                rows.append(new_row)

            tmp_path = csv_path + ".tmp"
            with open(tmp_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow(header)
                writer.writerows(rows)
            os.replace(tmp_path, csv_path)

            state._history_row_cache[cache_key] = new_row

        except Exception as e:
            state.error = f"歷史寫入失敗 ({c}): {e}"


def trigger_sheet_write(state: StockState):
    with state._sheet_lock:
        if state._sheet_writing:
            return
        state._sheet_writing = True

    def _run():
        try:
            push_to_sheet(state)
        except Exception as e:
            state.error = f"歷史資料寫入執行緒失敗: {e}"
        finally:
            with state._sheet_lock:
                state._sheet_writing = False

    t = threading.Thread(target=_run, daemon=True)
    t.start()


# --- CSV 寫入功能 ----------------──

def push_to_csv(state: StockState):
    """將即時報價獨立儲存為單一 CSV 檔案"""
    today = now_tpe().strftime("%Y-%m-%d")

    # 在鎖保護下取出 snapshot，取完立即釋放，避免長時間持鎖阻塞 GUI 讀取
    with state._data_lock:
        data_snapshot = dict(state.data)

    try:
        rows = []
        headers = [
            "名稱/代碼", "時間", "當前價", "昨收", "漲跌", "漲跌幅(%)",
            "開盤", "最高", "最低", "總量(張)", "振幅(%)",
            "委買價1", "委買量1", "委買價2", "委買量2",
            "委買價3", "委買量3", "委買價4", "委買量4",
            "委買價5", "委買量5",
            "委賣價1", "委賣量1", "委賣價2", "委賣量2",
            "委賣價3", "委賣量3", "委賣價4", "委賣量4",
            "委賣價5", "委賣量5",
        ]

        def build_rt_row(label: str, d: dict) -> list:
            bid_cols = []
            ask_cols = []
            for idx in range(5):
                bid_cols += [d["bid_prices"][idx], d["bid_vols"][idx]]
                ask_cols += [d["ask_prices"][idx], d["ask_vols"][idx]]

            return [
                label,
                f"{today} {d['time']}",
                d["price"],
                d["prev_close"],
                d["chg"],
                f"{d['pct']}%",
                d["open"],
                d["high"],
                d["low"],
                d["volume"],
                f"{d['amplitude']}%",
                *bid_cols,
                *ask_cols,
            ]

        for idx_info in INDEX_LIST:
            d = data_snapshot.get(idx_info["code"])
            if d and d["price"] > 0:
                rows.append(build_rt_row(idx_info["name"], d))

        for c in WATCH_LIST:
            d = data_snapshot.get(c)
            if d and d["price"] > 0:
                label = f"{d['name']}({c})" if d["name"] else c
                rows.append(build_rt_row(label, d))

        for c in sorted(list(ETF_CODES)):
            d = data_snapshot.get(c)
            if d and d["price"] > 0:
                label = f"{d['name']}({c})" if d["name"] else c
                rows.append(build_rt_row(label, d))

        # 覆蓋寫入 CSV 檔案
        with open(CSV_FILE, mode="w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(rows)

    except Exception as e:
        state.error = f"CSV 寫入失敗: {e}"


def trigger_csv_write(state: StockState):
    with state._csv_lock:
        if state._csv_writing:
            return
        state._csv_writing = True

    def _run():
        try:
            push_to_csv(state)
        except Exception as e:
            state.error = f"CSV 執行緒失敗: {e}"
        finally:
            with state._csv_lock:
                state._csv_writing = False

    t = threading.Thread(target=_run, daemon=True)
    t.start()


# --- Tkinter UI 視窗實作 -----------------------

class StockApp:
    def __init__(self, root, state):
        self.root = root
        self.state = state
        
        self.root.title("即時報價監控系統")
        self.root.geometry("1100x800")
        self.root.configure(bg="#1e1e1e")
        
        # 使用等寬字型確保表格基本字形一致
        self.display_font = tkfont.Font(family="Courier New", size=11, weight="bold")
        self.header_font = tkfont.Font(family="Microsoft JhengHei", size=11, weight="bold")
        
        # 計算當前螢幕與字型配置下的基礎字寬（動態對齊關鍵）
        char_width = self.display_font.measure("0")
        
        # 頂部狀態列
        self.status_frame = tk.Frame(root, bg="#2d2d2d", padx=10, pady=5)
        self.status_frame.pack(fill=tk.X, side=tk.TOP)
        
        self.status_label = tk.Label(
            self.status_frame, 
            text="載入中...", 
            font=self.header_font, 
            bg="#2d2d2d", 
            fg="#ffffff", 
            justify=tk.LEFT, 
            anchor="w"
        )
        self.status_label.pack(fill=tk.X)
        
        # 主要內文顯示區（加上滾動條）
        self.main_frame = tk.Frame(root, bg="#1e1e1e")
        self.main_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        self.scrollbar = tk.Scrollbar(self.main_frame)
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 修正：改用動態計算的字寬單位設定 tabs，完美適應不同高解析度與 DPI 的電腦螢幕
        self.text_area = tk.Text(
            self.main_frame, 
            font=self.display_font, 
            bg="#1e1e1e", 
            fg="#ffffff", 
            yscrollcommand=self.scrollbar.set,
            wrap=tk.NONE,
            bd=0,
            highlightthickness=0,
            tabs=(
                char_width * 34, "right", 
                char_width * 46, "right", 
                char_width * 58, "right", 
                char_width * 70, "right", 
                char_width * 82, "right"
            )
        )
        self.text_area.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.scrollbar.config(command=self.text_area.yview)
        
        # 註冊顏色標籤樣式
        self.text_area.tag_config("red", foreground="#ff4d4d")
        self.text_area.tag_config("green", foreground="#2ed573")
        self.text_area.tag_config("white", foreground="#ffffff")
        self.text_area.tag_config("title", foreground="#ffd32a")
        self.text_area.tag_config("dim", foreground="#747d8c")
        self.text_area.tag_config("special", foreground="#ffffff", background="#3742fa")
        self.text_area.tag_config("blue", foreground="#3498db")
        self.text_area.tag_config("alert_up", foreground="#000000", background="#ffa502")
        
        # 將突發下跌警示的背景色改為深紫色（#8e44ad），文字維持白色
        self.text_area.tag_config("alert_down", foreground="#ffffff", background="#8e44ad")        
        
        self.text_area.tag_config("bg_red", foreground="#ffffff", background="#ff4d4d")
        self.text_area.tag_config("bg_green", foreground="#ffffff", background="#2ed573")

        # 啟動背景網路抓取與排程
        self.start_background_loop()
        self.update_gui_loop()

    def start_background_loop(self):
        """獨立執行緒：專職處理網路請求、每分鐘寫入即時 CSV 與每 10 分鐘寫入歷史 CSV"""
        def _loop():
            last_csv_time = 0.0
            last_excel_time = 0.0
            
            # 設定更新週期（秒）
            csv_interval = 60      # 每 1 分鐘更新即時報價 CSV
            excel_interval = 600   # 剩下的歷史分頁隔久一點，每 10 分鐘更新一次

        # 調整初次觸發機制，讓系統啟動時能立刻有第一筆初始快取，不需乾等 10 秒
            refresh(self.state)
            trigger_csv_write(self.state)
            trigger_sheet_write(self.state)
            last_csv_time = time.time()
            last_excel_time = time.time()

            while True:
                time.sleep(10)
                refresh(self.state)
                current_now = time.time()
                
                # 判斷是否觸發每分鐘 CSV 寫入
                if current_now - last_csv_time >= csv_interval:
                    trigger_csv_write(self.state)
                    last_csv_time = current_now
                    
                # 判斷是否觸發每 10 分鐘歷史資料寫入
                if current_now - last_excel_time >= excel_interval:
                    trigger_sheet_write(self.state)
                    last_excel_time = current_now
        
        t = threading.Thread(target=_loop, daemon=True)
        t.start()

    def update_gui_loop(self):
        """主執行緒：定期刷新視窗文字內容（不造成畫面卡頓）"""
        self.render_data()
        self.root.after(500, self.update_gui_loop)

    def append_row_text(self, item: dict):
        """利用定位點將欄位直接以定位符號隔開，不受字元補白誤差影響"""
        lbl_str = item['label']
        code = item.get("code")
        
        if item["price"] == 0:
            name_tag = "blue" if code in HIGHLIGHT_STOCKS else "white"
            self.text_area.insert(tk.END, f"{lbl_str}", name_tag)
            self.text_area.insert(tk.END, "\t—\t—\t—\t—\t—\n", "white")
            return

        p_str = f"{item['price']:.2f}"
        c_str = f"{item['chg']:+.2f}"
        pct_str = f"{item['pct']:+.2f}%"
        h_str = f"{item['high']:.2f}"
        l_str = f"{item['low']:.2f}"
        
        col = price_color_tag(item["chg"])
        pct = item["pct"]
        alert = item.get("alert", False)
        alert_dir = item.get("alert_dir", "")

        # 判斷整行背景色覆蓋規則（處理突發變動 1% 的分流，與原本大於等於 5% 的暴漲跌高亮）
        if alert:
            if alert_dir == "up":
                self.text_area.insert(tk.END, f"{lbl_str}\t{p_str}\t{c_str}\t{pct_str}\t{h_str}\t{l_str}\n", "alert_up")
            else:
                self.text_area.insert(tk.END, f"{lbl_str}\t{p_str}\t{c_str}\t{pct_str}\t{h_str}\t{l_str}\n", "alert_down")
            return

        if pct >= 5:
            self.text_area.insert(tk.END, f"{lbl_str}\t{p_str}\t{c_str}\t{pct_str}\t{h_str}\t{l_str}\n", "bg_red")
            return

        if pct <= -5:
            self.text_area.insert(tk.END, f"{lbl_str}\t{p_str}\t{c_str}\t{pct_str}\t{h_str}\t{l_str}\n", "bg_green")
            return

        # 處理高亮特定特殊股：名稱固定藍色，股價、漲跌、漲跌幅全部隨漲跌變動顏色
        if code in HIGHLIGHT_STOCKS:
            self.text_area.insert(tk.END, f"{lbl_str}", "blue")
            self.text_area.insert(tk.END, f"\t{p_str}\t{c_str}\t{pct_str}\t{h_str}\t{l_str}\n", col)
            return

        # 一般正常個股分段填入：前置名稱維持白色，股價、漲跌、漲跌幅隨漲跌變動顏色
        self.text_area.insert(tk.END, f"{lbl_str}", "white")
        self.text_area.insert(tk.END, f"\t{p_str}\t{c_str}\t{pct_str}\t{h_str}\t{l_str}\n", col)

    def render_data(self):
        """將最新的 state 資料渲染進視窗"""
        # 1. 更新上方狀態列
        writing_hint = ""
        if self.state._sheet_writing:
            writing_hint = " [歷史資料寫入中...]"
        elif self.state._csv_writing:
            writing_hint = " [CSV即時寫入中...]"
            
        err_msg = f"\n錯誤訊息: {self.state.error}" if self.state.error else ""
        bad_msg = f" ⚠ 找不到股號: {', '.join(self.state.bad_codes)}" if self.state.bad_codes else ""
        status_text = (
            f" 本機時間：{self.state.local_time}   |   "
            f"交易所時間：{self.state.twse_time or '—'}{writing_hint}{err_msg}{bad_msg}"
        )
        self.status_label.config(text=status_text)
        
        # 紀錄當前滾動條的垂直比例位置
        current_scrollbar_pos = self.text_area.yview()

        # 2. 重新填入下方表格
        self.text_area.config(state=tk.NORMAL)
        self.text_area.delete("1.0", tk.END)
        
        # 在鎖保護下取出 snapshot，取完立即釋放，避免長時間持鎖阻塞背景寫入執行緒
        with self.state._data_lock:
            data_snapshot = dict(self.state.data)

        # 分流解析狀態結構
        indices = []
        etfs    = []
        stocks_by_industry = {}

        for idx_info in INDEX_LIST:
            d = data_snapshot.get(idx_info["code"])
            if d:
                indices.append({
                    "label": idx_info["name"], "price": d["price"], "chg": d["chg"], "pct": d["pct"],
                    "high": d["high"], "low": d["low"],
                    "code": idx_info["code"], "alert": d.get("alert", False), "alert_dir": d.get("alert_dir", "")
                })

        for c in WATCH_LIST:
            d = data_snapshot.get(c)
            if not d: continue
            label = f"{d['name']} ({c})" if d["name"] else c
            item_data = {
                "label": label, "price": d["price"], "chg": d["chg"], "pct": d["pct"],
                "high": d["high"], "low": d["low"],
                "code": c, "alert": d.get("alert", False), "alert_dir": d.get("alert_dir", "")
            }
            industry = self.state.get_industry(c)
            if industry not in stocks_by_industry:
                stocks_by_industry[industry] = []
            stocks_by_industry[industry].append(item_data)

        for c in ETF_CODES:
            d = data_snapshot.get(c)
            if not d: continue
            label = f"{d['name']} ({c})" if d["name"] else c
            item_data = {
                "label": label, "price": d["price"], "chg": d["chg"], "pct": d["pct"],
                "high": d["high"], "low": d["low"],
                "code": c, "alert": d.get("alert", False), "alert_dir": d.get("alert_dir", "")
            }
            etfs.append(item_data)

        # 標題行同步使用定位符號，確保絕對對齊
        header_line = "股名/股號\t股價\t漲跌\t漲跌幅(%)\t最高\t最低\n"
        separator   = "──────────────────────────────────────────────────────────────────────────────────────────\n"

        # 輸出 ── 指數區 ──
        if indices:
            self.text_area.insert(tk.END, "\n === 指數區 ===\n", "title")
            self.text_area.insert(tk.END, header_line, "dim")
            self.text_area.insert(tk.END, separator, "dim")
            for item in sorted(indices, key=lambda x: x["pct"], reverse=True):
                self.append_row_text(item)

        # 輸出 ── ETF 區 ──
        if etfs:
            self.text_area.insert(tk.END, "\n === ETF 區 ===\n", "title")
            self.text_area.insert(tk.END, header_line, "dim")
            self.text_area.insert(tk.END, separator, "dim")
            for item in sorted(etfs, key=lambda x: x["pct"], reverse=True):
                self.append_row_text(item)

        # 輸出 ── 個股區 ──
        if stocks_by_industry:
            self.text_area.insert(tk.END, "\n === 個股區 ===\n", "title")
            self.text_area.insert(tk.END, header_line, "dim")
            self.text_area.insert(tk.END, separator, "dim")
            
            # 依造指定產業順序排序
            # 原始清單中"電機機械"缺少"業"字，此處不更動以符合原功能
            print_order = [
                "半導體業", "電腦及週邊設備業", "電子零組件業", "其他電子業",
                "通信網路業", "光電業", "電子通路業", "資訊服務業",
                "其他業", "金融保險業", "航運業", "電機機械", "汽車工業",
                "化學工業", "塑膠工業", "橡膠工業", "鋼鐵工業", "建材營造業",
                "水泥工業", "玻璃陶瓷", "造紙工業", "電器電纜", "紡織纖維",
                "食品工業", "生技醫療業", "綠能環保", "數位雲端", "運動休閒",
                "居家生活", "觀光餐旅", "貿易百貨業", "油電燃氣業"
            ]
            
            existing_industries = list(stocks_by_industry.keys())
            sorted_industries = [ind for ind in print_order if ind in existing_industries]
            for ind in sorted(existing_industries):
                if ind not in sorted_industries:
                    sorted_industries.append(ind)

            for ind in sorted_industries:
                items = stocks_by_industry.get(ind, [])
                if not items: continue
                self.text_area.insert(tk.END, f"\n [{ind}]\n", "dim")
                for item in sorted(items, key=lambda x: x["pct"], reverse=True):
                    self.append_row_text(item)

        self.text_area.insert(tk.END, "\n\n   (TWSE MIS 即時行情系統連線中)\n", "dim")
        self.text_area.config(state=tk.DISABLED)
        
        # 重新填入完成後，將滾動條拉回剛才記憶的位置，避免視角位移
        self.text_area.yview_moveto(current_scrollbar_pos[0])
        
        self.root.update_idletasks()


# --- 主程式進入點 ----------------------------

def main():
    # 呼叫 Windows 系統 API，強制通知 OS 此程式自主處理高 DPI 縮放（消除字體毛邊）
    try:
        import ctypes
        ctypes.windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        pass

    root = tk.Tk()
    state = StockState()
    app = StockApp(root, state)
    root.mainloop()


if __name__ == "__main__":
    main()